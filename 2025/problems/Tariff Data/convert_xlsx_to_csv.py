import csv
from pathlib import Path

try:
    import win32com.client as win32
except ImportError:
    win32 = None

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl is not installed. Install it with: pip install openpyxl")
    raise


def xlsx_to_csv(xlsx_path: Path, output_dir: Path) -> None:
    """Convert one .xlsx file to one or multiple .csv files (one per sheet)."""
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    base_name = xlsx_path.stem

    sheets = list(wb.worksheets)
    multi_sheet = len(sheets) > 1

    for sheet in sheets:
        safe_title = "".join(
            c if c.isalnum() or c in ("-", "_") else "_" for c in sheet.title
        )

        if multi_sheet:
            csv_name = f"{base_name}__{safe_title}.csv"
        else:
            csv_name = f"{base_name}.csv"

        csv_path = output_dir / csv_name

        with csv_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(["" if cell is None else cell for cell in row])

        print(f"Saved: {csv_path}")


def xlsx_to_csv_via_excel(xlsx_path: Path, output_dir: Path) -> None:
    if win32 is None:
        raise RuntimeError(
            "pywin32 is not installed. Install it with: uv add pywin32 (or pip install pywin32)."
        )

    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    try:
        wb = excel.Workbooks.Open(str(xlsx_path))
        try:
            sheets = wb.Worksheets
            multi_sheet = sheets.Count > 1

            for sheet in sheets:
                safe_title = "".join(
                    c if c.isalnum() or c in ("-", "_") else "_" for c in sheet.Name
                )

                if multi_sheet:
                    csv_name = f"{xlsx_path.stem}__{safe_title}.csv"
                else:
                    csv_name = f"{xlsx_path.stem}.csv"

                csv_path = output_dir / csv_name
                sheet.SaveAs(str(csv_path), FileFormat=6)
                print(f"Saved via Excel: {csv_path}")
        finally:
            wb.Close(SaveChanges=False)
    finally:
        excel.Quit()


def should_use_excel_for(path: Path) -> bool:
    return path.name.startswith("DataWeb-Query-")


def convert_tree(root: Path) -> None:
    """Recursively convert all .xlsx files under root to .csv in-place."""
    for path in root.rglob("*.xlsx"):
        if path.name.startswith("~$"):
            continue
        if should_use_excel_for(path):
            xlsx_to_csv_via_excel(path, path.parent)
        else:
            xlsx_to_csv(path, path.parent)


if __name__ == "__main__":
    root_dir = Path(__file__).parent
    print(f"Converting all .xlsx files under: {root_dir}")
    convert_tree(root_dir)
    print("Done.")
